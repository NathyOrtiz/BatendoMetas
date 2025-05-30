#Importa os módulos necessários para o Flask funcionar
from flask import Flask, render_template, request, redirect
#Importando as funções da biblioteca openpyxl para criar e manipular
#um arquivo excel
from openpyxl import Workbook,load_workbook
#Biblioteca para verificar a existência de um arquivo excel
import os
#Criar de fato a nossa aplicação
app=Flask(__name__)
#Definindo o nome do arquivo da planilha excel
ARQUIVO='vendas.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()#Criando um arquivo Excel
    ws = wb.active #Selecionando a planilha ativa do projeto
    #Criando um cabeçalho para a planilha
    ws.append(["Nome","Vendas","Metas"])
    #Salvando o arquivo
    wb.save(ARQUIVO)

#Rota principal do site (formulário do cadastro de vendas)
@app.route('/')
def index():
    return render_template('index.html') #Retorna o conteúdo do arquivo index.html

#Rota que processa os dados do formulário e salva no excel
@app.route('/salvar',methods=['POST'])
def salvar():
    #Captura os dados de cada uma das caixas do formulário e atribui para as variáveis
    nome = request.form['nome']
    vendas = float(request.form['vendas'])
    meta = float(request.form['meta'])
#Abrindo o arquivo excel
    wb = load_workbook(ARQUIVO)
#Selecionando a planilha ativa - 1ª aba por padrão
    ws = wb.active

#Adiciona uma nova linha como lista com as informações do formulário
    ws.append([nome,vendas,meta])

#Salvando o arquivo excel
    wb.save(ARQUIVO)

#Redirecionando a rota para /analisar (onde abrirá uma nova página)
#passando por parâmetro o nome do funcionário.
    return redirect('/analisar?nome='+nome)

#Rota que analisa se o funcionário bateu a meta e calcula o bônus
@app.route('/analisar')
def analisar():
#Pega o nome do funcionário enviado como parâmetro pela URL
    nome_param = request.args.get('nome')
    wb = load_workbook(ARQUIVO)
    ws = wb.active
#Loop for: percorre todas as linhas da planilha a partir da 2ª linha
#(pois a 1ª linha é o cabeçalho) values_only retorna apenas os valores
    for linha in ws.iter_rows(min_row=2, values_only=True):
        nome,vendas,meta = linha
#A variável linha sempre receberá três valores, ex:
#linha=('Ana', 45,50) na linha de codigo acima, estamos atribuindo cada
#elemento da linha a uma variável na sequencia
#as variáveis nome,vendas,meta recebem Ana, 45, 50 respectivamente.
#Nome técnico desse processo é desempacotamento.

#Verifica se o nome atual da linha é o mesmo enviado na URL
    if nome == nome_param:
        meta_batida = (vendas >= meta)
#Se a meta for batida bonus recebe o resultado do cálculo de 15% das vendas
#caso contrário receberá 0
        bonus = round(vendas*0.15,2) if meta_batida else 0
#Exibir a tela Resultado com as informações dos cálculos
        return render_template('resultado.html',
                               nome = nome,meta_batida = meta_batida,
                               bonus = bonus)
    return "Funcionário não encontrado"

#Rota que mostra a página do histórico de todos os funcionários cadastrados
@app.route('/historico')
def historico():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
#Converte os dados da planilha (a partir da 2ª linha) em uma tupla
    dados = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('historico.html', dados = dados)
#Iniciando o Flask no modo desenvolvedor Debug
if __name__ == '__main__':
    app.run(debug=True)
